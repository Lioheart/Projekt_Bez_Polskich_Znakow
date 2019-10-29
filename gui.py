# -*- coding: utf-8 -*-
# LOGOWANIE
import os
import sys

from PyQt5 import QtWidgets
from PyQt5.QtCore import QSortFilterProxyModel, Qt, pyqtSlot, QRegExp
from PyQt5.QtGui import QIcon, QPalette, QColor, QCursor
from PyQt5.QtSql import QSqlDatabase, QSqlTableModel
from PyQt5.QtWidgets import QApplication, QDesktopWidget, QPushButton, \
    QHBoxLayout, QVBoxLayout, QLineEdit, QGroupBox, QDialog, QFormLayout, \
    QLabel, QMainWindow, QAction, QInputDialog, QMessageBox, QFileDialog, \
    QWidget, QComboBox, QTableView, QAbstractItemView, QMenu

from baza import polaczenie
from opcje_qt import Wewnatrz
from uzytkownicy import opcje_uzytkownik, dodaj_uzytkownik

id_user = 0


# centrowanie
def center(self):
    qr = self.frameGeometry()
    cp = QDesktopWidget().availableGeometry().center()
    qr.moveCenter(cp)
    self.move(qr.topLeft())


class Logowanie(QDialog):

    def __init__(self):
        super().__init__()

        # Edycja podwietlenia
        paleta = self.palette()
        paleta.setColor(QPalette.Highlight, QColor(233, 107, 57))
        self.setPalette(paleta)
        # Usuwa ramkę
        # self.setWindowFlags(Qt.FramelessWindowHint)

        # Dane okna formularza
        self.haslo = QLineEdit()
        self.haslo.setEchoMode(QLineEdit.Password)
        self.login = QLineEdit()
        self.formularz = QGroupBox("Logowanie")
        self.title = 'Logowanie'
        self.setWhatsThis('Pole logowania')
        self.width = 300
        self.height = 150
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Logowanie do wykazu narzędzi')
        self.setWindowIcon(QIcon('icons/cow.png'))
        self.setFixedSize(self.width, self.height)

        ok_button = QPushButton("OK")
        cancel_button = QPushButton("Anuluj")
        # Box odpowiedzialny za OK i Anuluj
        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(ok_button)
        hbox.addWidget(cancel_button)

        # layout logowania
        layout = QFormLayout()
        layout.addRow(QLabel("Login:"), self.login)
        layout.addRow(QLabel("Hasło:"), self.haslo)
        self.formularz.setLayout(layout)

        # Do USUNIECIA
        # self.login.setText('admin')
        # self.haslo.setText('admin')

        login = str(self.login.text())
        self.plik = os.path.isfile(os.environ['LOCALAPPDATA'] + '\\pbpz.cfg')
        if self.plik:
            file = open(os.environ['LOCALAPPDATA'] + '\\pbpz.cfg', 'r+', -1,
                        'utf-8')
            self.login.setText(file.read())
            file.write(login)
            file.close()

        center(self)

        # connecty
        ok_button.clicked.connect(self.logowanie)
        cancel_button.clicked.connect(self.reject)

        main_layout = QVBoxLayout()
        main_layout.addWidget(self.formularz)
        main_layout.addLayout(hbox)
        self.setLayout(main_layout)

        self.show()
        self.haslo.setFocus()

    def logowanie(self):
        login = str(self.login.text())
        query = "SELECT iduzytkownicy FROM uzytkownicy WHERE nazwa_uz = '" + \
                login + "' AND haslo_uz = '" + str(self.haslo.text()) + "';"
        global id_user
        id_user = polaczenie(query)
        if id_user:
            print("Logowanie udane")
            self.accept()
        else:
            print("Coś jest nie tak...")
            QMessageBox.information(self, 'Błąd logowania', "Niepoprawne "
                                                            "hasło lub login"
                                                            " \nSpróbuj "
                                                            "ponownie",
                                    QMessageBox.Ok, QMessageBox.Ok)
        if not self.plik:
            file = open(os.environ['LOCALAPPDATA'] + '\\pbpz.cfg', 'x', -1,
                        'utf-8')
            file.write(login)
            file.close()


class Wyswietl(QWidget):
    def __init__(self, parent):
        super(QWidget, self).__init__(parent)
        self.parent = parent
        self.proxy = QSortFilterProxyModel(self)
        self.edit_wysz = QLineEdit(self)
        self.combo_typ = QComboBox(self)
        self.lbl_wysz = QLabel("Wyszukaj")
        self.lbl_typ = QLabel("Wybierz typ narzędzia:")
        self.table = QTableView(self)
        db = QSqlDatabase.addDatabase('QSQLITE')
        db.setDatabaseName('poo.db')
        if db.open():
            print('Otworzono bazę danych')
        self.model = QSqlTableModel(self, db)

        self.formularz = QGroupBox("Wyświetl narzędzia")
        self.initUI()

    def initUI(self):
        typy_narzedzi = [
            'Brak',
            'Frez palcowy',
            'Frez płytkowy (głowica)',
            'Gwintownik',
            'Nóż tokarski',
            # 'Oprawka',
            'Piła',
            'Pozostałe',
            'Rozwiertak',
            'Wiertło',
            'Wiertło składane',
            'Wygniatak'
        ]
        self.widok()
        self.combo_typ.addItems(typy_narzedzi)

        cancel_button = QPushButton("Cofnij")
        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(cancel_button)

        l_h = QHBoxLayout()
        l_h.addWidget(self.lbl_typ)
        l_h.addWidget(self.combo_typ)

        l_h2 = QHBoxLayout()
        l_h2.addWidget(self.lbl_wysz)
        l_h2.addWidget(self.edit_wysz)

        layout_v = QVBoxLayout()
        layout_v.addLayout(l_h)
        layout_v.addLayout(l_h2)
        layout_v.addWidget(self.table)

        self.proxy.setSourceModel(self.model)
        self.table.setModel(self.proxy)

        main_layout = QVBoxLayout()
        self.formularz.setLayout(layout_v)
        main_layout.addWidget(self.formularz)
        main_layout.addLayout(hbox)
        self.setLayout(main_layout)

        cancel_button.clicked.connect(self.anulowanie)
        self.edit_wysz.textChanged.connect(self.wyszukiwanie)
        self.combo_typ.activated[str].connect(self.onActiveNarz)
        # Menu kontekstowe własne
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.prawoklik)

    def prawoklik(self):
        menu = QMenu(self)
        if self.model.rowCount():
            akcja = QAction('Usuń narzędzie', self)
            akcja.triggered.connect(self.usun_wiersz)
            menu.addAction(akcja)
            menu.exec_(QCursor.pos())

    def usun_wiersz(self):
        ok = QMessageBox.question(self, 'Potwierdzenie',
                                  'Czy na pewno chcesz usunąć narzędzie?',
                                  QMessageBox.Ok, QMessageBox.Cancel)
        if ok == QMessageBox.Ok:
            selected = self.table.currentIndex()
            self.model.removeRow(selected.row())
            self.model.submitAll()
            self.model.select()
            self.parent.statusBar().showMessage("Usunięto narzędzie", 10000)

    def onActiveNarz(self, tekst):
        slownik = {
            'Frez palcowy': 'frezy_palcowe',
            'Frez płytkowy (głowica)': 'frezy_plytkowe',
            'Gwintownik': 'gwintowniki',
            'Nóż tokarski': 'noze_tokarskie',
            'Oprawka': 'oprawki',
            'Piła': 'pily',
            'Pozostałe': 'pozostale',
            'Rozwiertak': 'rozwiertaki',
            'Wiertło': 'wiertla',
            'Wiertło składane': 'wiertla_skladane',
            'Wygniatak': 'wygniataki',
            'Brak': ''
        }
        naglowki = {
            'idfrezy_palcowe': 'ID',
            'symbol_freza': 'Symbol',
            'producent_fr': 'Producent',
            'srednica_fr': 'Średnica',
            'dl_fr': 'Długość całkowita',
            'dl_rob_fr': 'Długość robocza',
            'idfrezy_plytkowe': 'ID',
            'symbol_frez_pl': 'Symbol',
            'producent_fp': 'Producent',
            'srednica_fr_pl': 'Średnica',
            'ilosc_plytek': 'Ilość płytek',
            'symbol_pl': 'Symbol płytek',
            'ilosc_krawedzi_pl': 'Ilość krawędzi płytki',
            'idgwintowniki': 'ID',
            'symbol_g': 'Symbol',
            'producent_gw': 'Producent',
            'rozmiar_gwintu': 'Rozmiar gwintu i skok',
            'typ_gwintownika': 'Typ gwintownika',
            'idnoze_tokarskie': 'ID',
            'symbol_n': 'Symbol',
            'producent_n': 'Producent',
            'plytki_n': 'Symbol płytek',
            'ilosc_krawedzi_pl_n': 'Ilość krawędzi',
            'idpily': 'ID',
            'symbol_p': 'Symbol',
            'producent_pil': 'Producent',
            'srednica_p': 'Średnica',
            'grubosc_p': 'Grubość',
            'rodzaj_pl_p': 'Symbol płytek',
            'ilosc_pl_p': 'Ilość płytek',
            'ilosc_kraw_p': 'Ilość krawędzi płytki',
            'idpozostale': 'ID',
            'symbol_poz': 'Symbol',
            'producent_poz': 'Producent',
            'srednica_poz': 'Średnica',
            'ilosc_pl_poz': 'Ilość płytek',
            'plytki_poz': 'Symbol płytek',
            'idrozwiertaki': 'ID',
            'symbol_r': 'Symbol',
            'producent_roz': 'Producent',
            'rozmiar_r': 'Rozmiar',
            'idwiertla': 'ID',
            'symbol_w': 'Symbol',
            'producent_w': 'Producent',
            'srednica_w': 'Średnica',
            'dlugosc_w': 'Długość [mm]',
            'idwiertla_skladane': 'ID',
            'symbol_w_skl': 'Symbol',
            'producent_ws': 'Producent',
            'srednica_w_skl': 'Średnica',
            'plytki_w_skl': 'Symbol płytek',
            'idwygniataki': 'ID',
            'symbol_wyg': 'Symbol',
            'producent_wyg': 'Producent',
            'rozmiar_gw': 'Rozmiar gwintu'
        }
        self.model.setTable(slownik[tekst])
        self.model.setEditStrategy(QSqlTableModel.OnManualSubmit)
        self.model.select()

        # Ustawianie nagłówków
        ilosc_kolumn = self.model.columnCount()
        for i in range(ilosc_kolumn):
            nazwa_kolumn = self.model.headerData(i, Qt.Horizontal)
            self.model.setHeaderData(i, Qt.Horizontal, naglowki[nazwa_kolumn])

    @pyqtSlot(str)
    def wyszukiwanie(self, text):
        search = QRegExp(text,
                         Qt.CaseInsensitive,
                         QRegExp.RegExp
                         )
        self.proxy.setFilterRegExp(search)
        # Odpowiedzialne za kolumnę, po której filtruje
        self.proxy.setFilterKeyColumn(-1)

    def widok(self):
        # Ustawianie własciwości widoku
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.setModel(self.model)

    def anulowanie(self):
        self.parent.statusBar().clearMessage()
        from opcje_qt import Wewnatrz
        menu_gl = Wewnatrz(self.parent)
        self.parent.setCentralWidget(menu_gl)


class Window(QMainWindow):

    def __init__(self, user):
        super().__init__()
        self.id_user = user
        self.widget = Wewnatrz(self)
        self.title = 'Wykaz narzędzi'
        self.width = 1440
        self.height = 900

        # Edycja podwietlenia głównego okna
        paleta = self.palette()
        paleta.setColor(QPalette.Highlight, QColor(233, 107, 57))
        self.setPalette(paleta)

        self.initUI()

    def initUI(self):
        normy_lista = (
            'Normy',
            'SELECT * FROM detale',
            'Baza nie zawiera żadnych norm. Plik nie zostanie zapisany.',
        )

        self.setWindowTitle(self.title)
        self.setWindowIcon(QIcon('icons/cow.png'))
        self.resize(self.width, self.height)
        center(self)
        self.setCentralWidget(self.widget)  # centrowanie widżetu na pełneokno

        # Wyswietlanie
        # todo zaproponowanie skrótów
        wysw_act1 = QAction('Wyświetl normy pozycji', self)
        # wysw_act1.setShortcut('Ctrl+Q')
        wysw_act1.triggered.connect(self.norma)
        wysw_act2 = QAction('Wyświetl narzędzia pozycji', self)
        # wysw_act2.setShortcut('Ctrl+Q')
        wysw_act2.triggered.connect(self.wyswietl)

        # Wydruk
        wydr_act1 = QAction('Wydrukuj normy do pliku', self)
        # wydr_act1.setShortcut('Ctrl+Q')
        wydr_act1.triggered.connect(
            lambda checked, normyl=normy_lista: self.wybor_norma(
                normyl)
        )
        wydr_act2 = QAction('Wydrukuj wykaz narzędzi do pliku', self)
        # wydr_act2.setShortcut('Ctrl+Q')
        # wydr_act2.triggered.connect(
        #     lambda checked, pozl=None: self.wybor(pozl)
        # )
        wydr_act2.triggered.connect(self.wybor)

        # Opcje
        opcje_act1 = QAction('Zmiana hasła', self)
        # wydr_act1.setShortcut('Ctrl+Q')
        opcje_act1.triggered.connect(self.uzytkownik)
        opcje_act2 = QAction('Dodaj użytkownika', self)
        opcje_act2.triggered.connect(self.nowy_uzytkownik)
        opcje_act3 = QAction('Informacje o autorze', self)
        # opcje_act2.setShortcut('Ctrl+Q')
        opcje_act3.triggered.connect(self.about)

        menubar = self.menuBar()
        wyswietlanie = menubar.addMenu('Wyświetl')
        wydrukowanie = menubar.addMenu('Wydrukuj')
        opcje = menubar.addMenu('Opcje')
        wyswietlanie.addAction(wysw_act1)
        wyswietlanie.addAction(wysw_act2)
        wydrukowanie.addAction(wydr_act1)
        wydrukowanie.addAction(wydr_act2)
        opcje.addAction(opcje_act1)
        if id_user[0] == 1:
            opcje.addAction(opcje_act2)
        opcje.addAction(opcje_act3)

        self.show()

    def norma(self):
        from norma import Norma
        nor = Norma(self)
        self.setCentralWidget(nor)

    def wyswietl(self):
        wysw = Wyswietl(self)
        self.setCentralWidget(wysw)

    def nowy_uzytkownik(self):
        text, ok = QInputDialog.getText(self, 'Dodaj użytkownika',
                                        'Wprowadź nowego użytkownika:')
        if ok and text:
            dodaj_uzytkownik(text)
        self.statusBar().showMessage("Dodano nowego użytkownika", 10000)

    def uzytkownik(self):
        text, ok = QInputDialog.getText(self, 'Zmiana hasła',
                                        'Wprowadź nowe hasło:',
                                        QLineEdit.Password)
        if ok and text:
            opcje_uzytkownik(text, id_user)
            self.statusBar().showMessage("Zmieniono hasło", 10000)

    def about(self):
        self.window = QMainWindow()
        from o_mnie import Ui_O_mnie
        self.ui = Ui_O_mnie()
        self.ui.setupUi(self.window)
        self.window.show()

    # uzytk = Uzytkownik(self)
    # self.setCentralWidget(uzytk)

    def wybor(self):
        poz_lista = [
            'Wykaz narzędzi ',
            "SELECT * FROM ",
            'Baza nie zawiera żadnych pozycji. Plik nie zostanie zapisany.',
        ]

        lista_poz = []
        lista_poz.append('Brak')
        from narzedzia_poz import naglowki
        for i in naglowki():
            if "/" in i[0]:
                lista_poz.append(i[0])
        inp = QInputDialog(self)
        inp.setWhatsThis('Wybierz pozycję aby eksportować do pliku')
        inp.setLabelText('Wybierz pozycję:')
        inp.setWindowTitle('Pozycje')
        inp.setComboBoxItems(lista_poz)
        inp.setCancelButtonText('Anuluj')

        if inp.exec_() == QDialog.Accepted:
            print(inp.textValue())
            poz_lista[0] += inp.textValue()
            poz_lista[1] += "'" + inp.textValue() + "'"
            if inp.textValue() != 'Brak':
                self.export(poz_lista)
                self.statusBar().showMessage(
                    "Wyeksportowano do pliku", 10000)
            else:
                QMessageBox.critical(self, 'Wybierz pozycję',
                                     'Nie wybrano żadnej pozycji!',
                                     QMessageBox.Ok,
                                     QMessageBox.Ok)

    def wybor_norma(self, lista):
        item, ok = QInputDialog.getItem(self, 'Wybierz', 'Typ:',
                                        ['Odkuwki', 'Kołnierze'], 0, False)
        if ok and item:
            if item == 'Kołnierze':
                lista = list(lista)
                lista[1] = 'SELECT * FROM kolnierze'
                lista = tuple(lista)
        else:
            return

        self.export(lista)
        self.statusBar().showMessage(
            "Wyeksportowano do pliku", 10000)

    def export(self, lista_arg):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Zapisywanie jako",
            os.path.expanduser("~") + '/desktop/' + lista_arg[0].replace('/',
                                                                         '-'),
            "Skoroszyt programu Excel (*.xlsx)",
            options=options
        )
        if file_name:
            print(file_name)

            import sqlite3
            conn = sqlite3.connect('poo.db')
            cursor = conn.cursor()
            mysel = cursor.execute(lista_arg[1])
            dane = mysel.fetchall()
            if not dane:
                QMessageBox.warning(self, 'Pusta zawartość',
                                    lista_arg[2],
                                    QMessageBox.Ok,
                                    QMessageBox.Ok)
            else:
                from xlsxwriter import Workbook
                workbook = Workbook(file_name)
                worksheet = workbook.add_worksheet()
                for i, row in enumerate(dane):
                    for j, value in enumerate(row):
                        if j == 0:
                            continue
                        worksheet.write(i, j, value)
                workbook.close()

                if 'Normy' in lista_arg:
                    stylizacja(file_name)
                else:
                    styl_pozycje(file_name, lista_arg[0][-7:])


def styl_pozycje(plik, nazwa='999/999'):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    from openpyxl.styles import Border
    from openpyxl.styles import Side
    from openpyxl.styles import Alignment
    from openpyxl.styles import PatternFill
    import openpyxl
    from openpyxl.drawing.image import Image
    nazwa_font = 'FL Pismo Techniczne'
    szer = 0.71
    szer_lista = [
        7,
        22.57,
        7.43,
        28.43,
        9.43,
        11,
        11,
        11
    ]
    merge_list = [
        'A1:C1',
        'D1:H1',
        'A2:C2',
        'D2:H2',
        'A3:C3',
        'D3:H3',
        'A4:A5',
        'B4:B5',
        'C4:D5',
        'E4:E5',
        'F4:F5',
        'G4:G5',
        'H4:H5',
        'A33:C33',
        'A34:C34',
        'A35:C35',
        'D33:E33',
        'D34:E34',
        'D35:E35',
        'F33:H33',
        'F34:H34',
        'F35:H35'
    ]
    for i in range(27):
        tekst = 'C' + str(i + 6) + ':D' + str(i + 6)
        merge_list.append(tekst)
    wys_lista = [
        62,
        12.75,
        23.25,
        12.75,
        12.75,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        25.5,
        12.75,
        12.75,
        12.75,
        12.75,
        12.75
    ]
    font_lista = [
        Font(size=18, bold=True, name=nazwa_font),
        Font(size=10, name=nazwa_font),
        Font(size=10, bold=True, name=nazwa_font)
    ]

    aligment = Alignment(horizontal='center', vertical='center',
                         wrap_text=True)
    teks_slownik = {
        'SPIS POMOCY WARSZTATOWYCH': [1, 4, font_lista[0]],
        'Skrócony kod detalu': [2, 1, font_lista[1]],
        'Nazwa detalu': [2, 4, font_lista[1]],
        nazwa: [3, 1, font_lista[0]],
        'Nr operacji': [4, 1, font_lista[2]],
        'Oprawki / Noże tokarskie': [4, 2, font_lista[2]],
        'Narzędzie': [4, 3, font_lista[2]],
        'Prędkość Vc': [4, 5, font_lista[2]],
        'Obroty n': [4, 6, font_lista[2]],
        'Posuw na ząb fz': [4, 7, font_lista[2]],
        'Posuw f': [4, 8, font_lista[2]],
        'Opracował, data, podpis': [33, 1, font_lista[1]],
        'Sprawdził, data, podpis': [33, 4, font_lista[1]],
        'Zatwierdził, data, podpis': [33, 6, font_lista[1]]
    }

    work = load_workbook(plik)
    ws = work[work.sheetnames[0]]
    for i in range(5):
        ws.insert_rows(0)
    ws.insert_cols(0)
    ws.insert_cols(4)
    for i, wys in enumerate(wys_lista):
        ws.row_dimensions[i + 1].height = wys
    for i, col in enumerate(szer_lista):
        ws.column_dimensions[get_column_letter(i + 1)].width = col + szer
    for val in merge_list:
        ws.merge_cells(val)

    czcionka = Font(name=nazwa_font)
    thin = Side(border_style="thin", color="000000")
    for i in range(8):
        for j in range(35):
            ws.cell(row=j + 1, column=i + 1).font = czcionka
            ws.cell(row=j + 1, column=i + 1).alignment = aligment
            if j >= 33:
                continue
            ws.cell(row=j + 1, column=i + 1).border = Border(
                thin, thin, thin, thin)

    for key, value in teks_slownik.items():
        ws.cell(value[0], value[1]).value = key
        ws.cell(value[0], value[1]).font = value[2]
        ws.cell(value[0], value[1]).alignment = aligment

    wypelnienie = PatternFill(start_color='EEECE1', end_color='EEECE1',
                              fill_type='solid')
    for i in range(8):
        for j in range(26):
            if not (j % 2 == 0):
                ws.cell(row=j + 6, column=i + 1).fill = wypelnienie

    img = openpyxl.drawing.image.Image('logo.png')
    # 100 = 2,65cm
    img.height = 80.75471698
    img.width = 269.43396226
    ws.add_image(img, 'A1')

    ws.oddFooter.left.text = 'P-13.01.00.08'
    ws.oddFooter.left.size = 12
    ws.oddFooter.left.font = nazwa_font
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = True
    # 1cm = 0.4
    ws.page_margins.left = 0.24
    ws.page_margins.right = 0.24
    ws.page_margins.top = 0.76
    ws.page_margins.bottom = 0.76
    ws.page_margins.header = 0.32
    ws.page_margins.footer = 0.32
    ws.page_setup.print_area = 'A1:H37'
    work.save(plik)
    work.close()
    print('Wykonano')


def stylizacja(plik):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    lista = [
        'Detal',
        'Maszyna',
        'Ilość maszyn',
        'Ilość raportowanych sztuk',
        'Numer operacji',
        'Norma',
        'Uwagi'
    ]

    szer = 0.71
    szer_lista = [
        7.14,
        11.86,
        11.29,
        23.43,
        14.14,
        6.29,
        20
    ]
    work = load_workbook(plik)
    worksheet = work[work.sheetnames[0]]
    worksheet.insert_rows(0)
    worksheet.delete_cols(12)
    worksheet.delete_cols(7, 3)
    worksheet.delete_cols(1)

    for h in range(7):
        worksheet.column_dimensions[get_column_letter(h + 1)].width = \
            szer_lista[h] + szer

    for i, col in enumerate(lista):
        worksheet.cell(row=1, column=i + 1).value = col

    # tabela
    rozmiar = 'A1:G' + str(len(worksheet['A']))
    tab = Table(displayName='Table1', ref=rozmiar)
    style = TableStyleInfo(
        name='TableStyleMedium1',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    worksheet.add_table(tab)

    work.save(plik)
    work.close()
    print('Udane')


def aplikacja():
    app = QApplication(sys.argv)
    from PyQt5.QtWidgets import QStyleFactory
    app.setStyle(QStyleFactory.create('Fusion'))
    ex = Logowanie()

    if ex.exec_() == QtWidgets.QDialog.Accepted:
        main_window = Window(id_user)
        # main_window.showMaximized()  # maksymalizuje
        main_window.show()
        sys.exit(app.exec_())
